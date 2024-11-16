import os
import arcpy
from arcpy.sa import *
import pandas as pd
from openpyxl import Workbook, load_workbook

workspace = 'D:/Desktop/cfb/data/'
# test environment:
# arcpy.env.workspace = r"D:\Desktop\cfb\data"
# arcpy.env.overwriteOutput = True
# arcpy.env.outputCoordinateSystem = arcpy.SpatialReference(4548)


class Simplify:
    def __init__(self):
        self.input_raster = None
        self.output_raster = None
        self.resample_cell_size = None
        self.filter_cell_size = None

    def simplify_params_set(self, obj):
        self.input_raster = obj['input_raster']
        self.output_raster = obj['output_raster']
        self.resample_cell_size = obj['resample_cell_size']
        self.filter_cell_size = obj['filter_cell_size']

    def simplify(self):
        arcpy.management.Resample(self.input_raster, "cache_resampled.tif", self.resample_cell_size, "MAJORITY")
        neighborhood = NbrRectangle(self.filter_cell_size, self.filter_cell_size, "CELL")
        majority_raster = FocalStatistics("cache_resampled.tif", neighborhood, "MAJORITY")
        SetNull(majority_raster, majority_raster, "VALUE = 0").save(self.output_raster)
        arcpy.management.Delete("cache_resampled.tif")


class NeighborRecognition:
    def __init__(self):
        self.input_raster = None
        self.output_feature = None
        self.forest_index = None

    def neighbor_recognition_params_set(self, obj):
        self.input_raster = obj['simplified_raster']
        self.output_feature = obj['output_feature']
        self.forest_index = obj['forest_index']

    def __extract_extent(self):
        extract_forest = SetNull(self.input_raster, self.input_raster, f"VALUE <> {self.forest_index}")
        extract_lulc = SetNull(self.input_raster, self.input_raster, f"VALUE = {self.forest_index}")
        arcpy.conversion.RasterToPolygon(extract_forest, "forest_extent.shp", "SIMPLIFY")
        arcpy.conversion.RasterToPolygon(extract_lulc, "lulc_extent.shp", "SIMPLIFY")
        arcpy.management.AddField("forest_extent.shp", "SHAPE_Area", "DOUBLE")
        arcpy.management.CalculateGeometryAttributes("forest_extent.shp", "SHAPE_Area AREA", area_unit="SQUARE_METERS")
        arcpy.management.MakeFeatureLayer("forest_extent.shp", "cache_lyr")
        arcpy.management.SelectLayerByAttribute("cache_lyr", "NEW_SELECTION", "SHAPE_Area < 300000000")
        arcpy.management.DeleteFeatures("cache_lyr")
        arcpy.management.Delete("cache_lyr")  # 显式释放内存

    def neighbor_recognition(self):
        self.__extract_extent()
        arcpy.analysis.Buffer("forest_extent.shp", "forest_extent_outward_buffer.shp", "1400 Meters", "OUTSIDE_ONLY")
        arcpy.analysis.Clip("lulc_extent.shp", "forest_extent_outward_buffer.shp", self.output_feature)
        arcpy.management.Delete("forest_extent_outward_buffer.shp")
        arcpy.management.Delete("lulc_extent.shp")


class EdgeEffectMeasure:
    def __init__(self):
        self.carbon_raster = None
        self.neighbor_extent = None
        self.output_table = None
        self.edge_distance = None
        self.distance_class = None

    def edge_effect_measure_params_set(self, obj):
        self.carbon_raster = obj['carbon_raster']
        self.neighbor_extent = obj['neighbor_extent']
        self.output_table = obj['output_table']
        self.edge_distance = obj['edge_distance']
        self.distance_class = obj['distance_class']

    def __carbon_points_extract(self):
        try:
            forest_carbon = ExtractByMask(self.carbon_raster, "forest_extent.shp")
            arcpy.management.CreateRandomPoints("D:/Desktop/cfb/data", "cache_carbon_points", constraining_extent="forest_extent.shp", number_of_points_or_field=2000, minimum_allowed_distance=5000)
            ExtractValuesToPoints("cache_carbon_points.shp", forest_carbon, "carbon_points.shp")
        finally:
            arcpy.management.Delete("cache_carbon_points.shp")

    def __accept_points(self):
        try:
            arcpy.management.MakeFeatureLayer("carbon_points.shp", "cache_point_lyr")
            arcpy.management.MakeFeatureLayer(self.neighbor_extent, "cache_polygon_lyr")
            arcpy.management.SelectLayerByAttribute("cache_point_lyr", "NEW_SELECTION", "RASTERVALU = -9999")
            arcpy.management.DeleteFeatures("cache_point_lyr")
            workbook = Workbook()
            if os.path.exists(f"{workspace}{self.output_table}"):
                os.remove(f"{workspace}{self.output_table}")
            workbook.save(f"{workspace}{self.output_table}")
            workbook = load_workbook(f"{workspace}{self.output_table}")
            step = self.edge_distance / self.distance_class
            with arcpy.da.SearchCursor(self.neighbor_extent, ['FID', 'gridcode']) as cursor:
                # 遍历每个邻域
                for row in cursor:
                    if f"{row[0]}_{row[1]}" not in workbook.sheetnames:
                        workbook.create_sheet(f"{row[0]}_{row[1]}")
                    sheet = workbook[f"{row[0]}_{row[1]}"]
                    sheet['A1'] = "distance"
                    sheet['B1'] = "value"
                    start_dis = 0
                    end_dis = start_dis + step
                    arcpy.management.SelectLayerByAttribute('cache_polygon_lyr', "NEW_SELECTION", f"FID = {row[0]}")
                    # 遍历每段邻接距离
                    for i in range(self.distance_class):
                        total = 0
                        count = 0
                        arcpy.management.SelectLayerByLocation("cache_point_lyr", "WITHIN_A_DISTANCE", 'cache_polygon_lyr', f'{end_dis} Kilometers', "NEW_SELECTION")
                        if start_dis != 0:
                            arcpy.management.SelectLayerByLocation("cache_point_lyr", "WITHIN_A_DISTANCE", 'cache_polygon_lyr', f'{start_dis} Kilometers', "REMOVE_FROM_SELECTION")
                        # 统计平均值
                        with arcpy.da.SearchCursor("cache_point_lyr", ['RASTERVALU']) as cursor2:
                            for row2 in cursor2:
                                total += row2[0]
                                count += 1
                        average = total / count if count != 0 else 0
                        # 写入表格
                        if average != 0:
                            sheet[f'A{i + 2}'] = end_dis
                            sheet[f'B{i + 2}'] = average
                            workbook.save(f"{workspace}{self.output_table}")
                        start_dis = end_dis
                        end_dis += step
        finally:
            arcpy.management.Delete("cache_point_lyr")
            arcpy.management.Delete("cache_polygon_lyr")

    def __gather_table(self):
        table = pd.ExcelFile(f"{workspace}{self.output_table}")
        gridcode_sheet_dict = {}
        for sheet in table.sheet_names:
            if not '_' in sheet:
                continue
            df_sheet = pd.read_excel(table, sheet_name=sheet)
            gridcode = sheet.split('_')[1]
            if gridcode not in gridcode_sheet_dict:
                gridcode_sheet_dict[gridcode] = []
            gridcode_sheet_dict[gridcode].append(df_sheet)
        summary_dict = {}
        for gridcode, sheet_list in gridcode_sheet_dict.items():
            combined_df = pd.concat(sheet_list)
            summary_df = combined_df.groupby('distance', as_index=False)['value'].mean()
            summary_dict[gridcode] = summary_df
        with engine.connect() as connection:
            for gridcode, summary_df in summary_dict.items():
                summary_df.to_sql(f'summary_{gridcode}', con=connection, schema='computed', if_exists='replace', index=False)

    def edge_effect_measure(self):
        self.__carbon_points_extract()
        self.__accept_points()
        self.__gather_table()
